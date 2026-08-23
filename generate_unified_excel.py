import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_master_excel():
    wb = openpyxl.Workbook()
    
    # Styles
    navy_fill = PatternFill(start_color="182B49", end_color="182B49", fill_type="solid")
    blue_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    light_green_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    light_blue_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
    zebra_fill = PatternFill(start_color="F8F9F9", end_color="F8F9F9", fill_type="solid")

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Arial", size=10, bold=True, color="000000")
    green_bold_font = Font(name="Arial", size=10, bold=True, color="1E8449")
    regular_font = Font(name="Arial", size=10, color="000000")
    
    thin_border = Border(
        left=Side(style='thin', color='BDC3C7'),
        right=Side(style='thin', color='BDC3C7'),
        top=Side(style='thin', color='BDC3C7'),
        bottom=Side(style='thin', color='BDC3C7')
    )

    # -------------------------------------------------------------------------
    # Sheet 1: Master Architectural Stages (E0 to Full WS-DBNet)
    # -------------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Master Summary (SFCD)"
    
    ws1_headers = [
        "Phase / Experiment", "Architecture Description", "S (Spatial)", "C (Context)", "FFM", "D (Decoder)",
        "Accuracy (%)", "Precision (%)", "Recall (%)", "F1 Score (%)", "IoU (%)", "Dice Score (%)", "mIoU (%)", "Test Loss", "Absolute mIoU Gain (%)"
    ]
    ws1.append(ws1_headers)
    
    ws1_rows = [
        ["Baseline (Phase E0)", "Original Dual-Branch DBCNet Baseline", "-", "-", "-", "-", 98.63, 94.88, 92.38, 93.61, 88.96, 93.61, 92.20, 0.0531, "Baseline"],
        ["Phase S1 (Spatial)", "CrossNet+ Multi-Scale Strip Convolutions (n=5,9,13)", "Yes", "-", "-", "-", 98.72, 95.04, 97.22, 95.98, 89.19, 95.98, 92.74, 0.0637, "+0.54%"],
        ["Phase SF Base", "Spatial Multi-Scale + 4-Stage FFM Base Model", "Yes", "-", "Yes", "-", 98.83, 95.26, 93.07, 94.15, 89.88, 94.15, 92.86, 0.0468, "+0.66%"],
        ["Phase C (Context Best)", "SF Base + Wavelet-Mamba Context (Sub 2.1)", "Yes", "Yes", "Yes", "-", 98.84, 92.88, 97.00, 94.38, 90.25, 94.38, 94.44, 0.0945, "+2.24%"],
        ["Phase D (Decoder Best)", "SF Base + Progressive CMM Decoder (Sub 4.1+4.3)", "Yes", "-", "Yes", "Yes", 98.80, 95.83, 97.19, 96.39, 93.03, 96.39, 93.31, 0.0601, "+1.11%"],
        ["Phase SFCD (Full WS-DBNet)", "Complete Proposed Architecture (S1 + F1 + C2.1 + D4.1_4.3)", "Yes", "Yes", "Yes", "Yes", 98.88, 96.51, 97.22, 96.86, 93.92, 96.86, 94.82, 0.0452, "+2.62% (Peak SOTA)"]
    ]
    for row in ws1_rows:
        ws1.append(row)

    # Format Sheet 1
    for col_idx in range(1, len(ws1_headers) + 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.fill = navy_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in range(2, len(ws1_rows) + 2):
        is_highlight = (row_idx == len(ws1_rows) + 1)
        for col_idx in range(1, len(ws1_headers) + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if is_highlight:
                cell.fill = light_green_fill
                cell.font = green_bold_font
            elif row_idx % 2 == 1:
                cell.fill = zebra_fill
                cell.font = regular_font
            else:
                cell.font = regular_font

    # -------------------------------------------------------------------------
    # Sheet 2: Section 2 Context Branch Powerset Matrix (Vishal's Results)
    # -------------------------------------------------------------------------
    ws2 = wb.create_sheet(title="Section 2 - Context Powerset")
    ws2_headers = [
        "Phase / Experiment", "Description", "Accuracy (%)", "Precision (%)", "Recall (%)", "F1 Score (%)", "IoU (%)", "Dice Score (%)", "mIoU (%)", "Test Loss"
    ]
    ws2.append(ws2_headers)
    
    ws2_rows = [
        ["Baseline (Phase E0)", "Original Dual-Branch DBCNet Baseline", 98.63, 94.88, 92.38, 93.61, 88.96, 93.61, 92.20, 0.0531],
        ["SF Base (Spatial + FFM)", "Spatial Multi-Scale + 4-Stage FFM Base Model", 98.83, 95.26, 93.07, 94.15, 89.88, 94.15, 92.86, 0.0468],
        ["sub-2-1-clean", "Sub 2.1 ON (Multi-Scale Strip Convs)", 98.84, 92.88, 97.00, 94.38, 90.25, 94.38, 94.44, 0.0945],
        ["sub-2-2-clean", "Sub 2.2 ON (ECA & Strip Pooling)", 98.69, 92.80, 95.61, 93.54, 88.91, 93.54, 93.70, 0.1022],
        ["sub-2-3-clean", "Sub 2.3 ON (Cross Bias & Spatial Gate)", 98.80, 93.06, 96.58, 94.22, 89.96, 94.22, 94.28, 0.1049],
        ["sub-2-1-2-2-clean", "Sub 2.1 + Sub 2.2 ON", 98.74, 92.04, 96.88, 93.46, 89.37, 93.46, 93.96, 0.1142],
        ["sub-2-1-2-3-clean", "Sub 2.1 + Sub 2.3 ON", 98.84, 93.30, 96.53, 94.35, 90.19, 94.35, 94.42, 0.0969],
        ["sub-2-2-2-3-clean", "Sub 2.2 + Sub 2.3 ON", 98.66, 91.68, 96.36, 93.28, 88.56, 93.28, 93.51, 0.1046],
        ["sub-2-all-clean", "Sub 2.1 + Sub 2.2 + Sub 2.3 ALL ON", 98.63, 92.22, 96.37, 93.61, 88.96, 93.61, 93.69, 0.1047]
    ]
    for row in ws2_rows:
        ws2.append(row)

    for col_idx in range(1, len(ws2_headers) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.fill = blue_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in range(2, len(ws2_rows) + 2):
        is_best = (row_idx == 4) # sub-2-1-clean
        for col_idx in range(1, len(ws2_headers) + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if is_best:
                cell.fill = light_green_fill
                cell.font = green_bold_font
            elif row_idx % 2 == 1:
                cell.fill = zebra_fill
                cell.font = regular_font
            else:
                cell.font = regular_font

    # -------------------------------------------------------------------------
    # Sheet 3: Section 4 Decoder Branch Powerset Matrix (Our Results)
    # -------------------------------------------------------------------------
    ws3 = wb.create_sheet(title="Section 4 - Decoder Powerset")
    ws3_headers = [
        "Phase / Experiment", "Description", "4.1 (5-Stage)", "4.2 (ECA Gate)", "4.3 (Prog Res)",
        "Accuracy (%)", "Precision (%)", "Recall (%)", "F1 Score (%)", "IoU (%)", "Dice Score (%)", "mIoU (%)", "Test Loss", "Absolute mIoU Gain (%)"
    ]
    ws3.append(ws3_headers)
    
    ws3_rows = [
        ["sub-4-1-clean", "Sub 4.1 ON (5-Stage CMM Coverage)", "Yes", "-", "-", 98.70, 92.52, 97.44, 93.92, 88.54, 93.92, 90.44, 0.0682, "Baseline CMM"],
        ["sub-4-2-clean", "Sub 4.2 ON (ECA Gating Swap)", "-", "Yes", "-", 98.68, 92.19, 97.83, 93.86, 88.43, 93.86, 90.29, 0.0691, "-0.15%"],
        ["sub-4-3-clean", "Sub 4.3 ON (Progressive Depthwise CMM)", "-", "-", "Yes", 98.73, 92.60, 97.85, 94.13, 88.91, 94.13, 90.78, 0.0664, "+0.34%"],
        ["sub-4-1-4-2-clean", "Sub 4.1 + Sub 4.2 ON", "Yes", "Yes", "-", 98.71, 92.71, 97.68, 94.08, 88.82, 94.08, 90.69, 0.0673, "+0.25%"],
        ["sub-4-1-4-3-clean (Optimal)", "Sub 4.1 + Sub 4.3 ON (5-Stage + Progressive CMM)", "Yes", "-", "Yes", 98.80, 95.83, 97.19, 96.39, 93.03, 96.39, 93.31, 0.0601, "+2.87% (Optimal)"],
        ["sub-4-2-4-3-clean", "Sub 4.2 + Sub 4.3 ON", "-", "Yes", "Yes", 98.66, 92.37, 97.45, 93.78, 88.29, 93.78, 90.13, 0.0705, "-0.31%"],
        ["sub-4-all-clean", "Sub 4.1 + Sub 4.2 + Sub 4.3 ALL ON", "Yes", "Yes", "Yes", 98.75, 93.36, 97.83, 94.52, 89.61, 94.52, 91.45, 0.0642, "+1.01%"]
    ]
    for row in ws3_rows:
        ws3.append(row)

    for col_idx in range(1, len(ws3_headers) + 1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.fill = navy_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in range(2, len(ws3_rows) + 2):
        is_opt = (row_idx == 6) # sub-4-1-4-3-clean
        for col_idx in range(1, len(ws3_headers) + 1):
            cell = ws3.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if is_opt:
                cell.fill = light_green_fill
                cell.font = green_bold_font
            elif row_idx % 2 == 1:
                cell.fill = zebra_fill
                cell.font = regular_font
            else:
                cell.font = regular_font

    # -------------------------------------------------------------------------
    # Sheet 4: Parameter & Reproducibility Verification
    # -------------------------------------------------------------------------
    ws4 = wb.create_sheet(title="Setup & Parameter Checklist")
    ws4_headers = ["Category", "Parameter Name", "Target / Paper Specification", "Our Implementation Value", "Verification Check"]
    ws4.append(ws4_headers)

    ws4_rows = [
        ["1. Dataset", "Dataset Name", "Glacial Lake Dataset", "Glacial Lake Dataset", "Verified Identical"],
        ["1. Dataset", "Data Split (Train/Val/Test)", "70% / 15% / 15%", "70% / 15% / 15% (1498 / 321 / 322)", "Verified Identical"],
        ["1. Dataset", "Image Resolution", "512 x 512", "512 x 512 (Normalized ImageNet)", "Verified Identical"],
        ["2. Training", "Epochs", "40 Epochs", "40 Epochs", "Verified Identical"],
        ["2. Training", "Batch Size", "2", "2", "Verified Identical"],
        ["2. Training", "Optimizer", "AdamW (lr=0.001, decay=1e-4)", "AdamW (lr=0.001, decay=1e-4)", "Verified Identical"],
        ["2. Training", "Scheduler", "Polynomial Decay + Warmup", "Polynomial Decay (power=0.9, 4 warmup ep)", "Verified Identical"],
        ["3. Loss", "Primary Loss Function", "BCE + Dice Loss", "BCE + Dice Loss (Standard Unmodified)", "Verified Identical"],
        ["4. Reproducibility", "Deterministic Random Seed", "3407", "torch.manual_seed(3407)", "Verified Identical"],
        ["4. Reproducibility", "Hardware Acceleration", "NVIDIA GPU / Mixed Precision", "RTX 2050 / PyTorch AMP FP16", "Verified Identical"]
    ]
    for row in ws4_rows:
        ws4.append(row)

    for col_idx in range(1, len(ws4_headers) + 1):
        cell = ws4.cell(row=1, column=col_idx)
        cell.fill = navy_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx in range(2, len(ws4_rows) + 2):
        for col_idx in range(1, len(ws4_headers) + 1):
            cell = ws4.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx == 5:
                cell.font = green_bold_font
            else:
                cell.font = regular_font

    # Auto-adjust column widths across all sheets
    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    out_excel = r"c:\Users\sm080\Downloads\glacial lake dataset\dbcnet_glacial_lakes\Glacial_Lake_Segmentation_Master_Ablation_Results.xlsx"
    wb.save(out_excel)
    print(f"Master Excel file successfully generated at: {out_excel}")

if __name__ == '__main__':
    create_master_excel()
